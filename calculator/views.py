from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook

from .serializers import (
    CalculationInputSerializer,
    CalculationOutputSerializer
)
from .services import OzonCalculator
from categories.models import Category


class CalculateAPIView(APIView):
    """
    API эндпоинт для расчета юнит-экономики товара на Ozon
    
    Принимает параметры товара и возвращает расчет прибыли
    для схем работы FBO и FBS
    """
    
    @extend_schema(
        request=CalculationInputSerializer,
        responses={200: CalculationOutputSerializer},
        description='Расчет юнит-экономики товара для схем FBO и FBS'
    )
    def post(self, request):
        """
        POST запрос для расчета прибыли
        """
        # Валидация входных данных
        input_serializer = CalculationInputSerializer(data=request.data)
        
        if not input_serializer.is_valid():
            return Response(
                {'errors': input_serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        validated_data = input_serializer.validated_data
        
        # Проверяем существование категории
        try:
            category = Category.objects.get(id=validated_data['category_id'])
        except Category.DoesNotExist:
            return Response(
                {'error': f'Категория с ID {validated_data["category_id"]} не найдена'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Создаем калькулятор и выполняем расчет
        try:
            calculator = OzonCalculator(
                category_id=validated_data['category_id'],
                price=validated_data['price'],
                weight=validated_data['weight'],
                volume=validated_data['volume'],
                tax_rate=validated_data['tax_rate'],
                buyout_rate=validated_data['buyout_rate'],
                delivery_time=validated_data['delivery_time'],
                ad_costs_rate=validated_data['ad_costs_rate'],
                cost_price=validated_data['cost_price'],
                other_costs=validated_data['other_costs'],
                monthly_sales=validated_data['monthly_sales']
            )
            
            # Выполняем расчет
            results = calculator.calculate_all()
            
            # Сериализуем результаты
            output_serializer = CalculationOutputSerializer(data=results)
            if output_serializer.is_valid():
                return Response(output_serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {'error': 'Ошибка при формировании результатов расчета'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        except Exception as e:
            return Response(
                {'error': f'Ошибка при расчете: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CalculateExportAPIView(APIView):
    """
    Экспорт результатов расчета юнит-экономики в Excel
    """

    @extend_schema(
        request=CalculationInputSerializer,
        responses={200: OpenApiTypes.BINARY},
        description='Расчет юнит-экономики и выгрузка в Excel'
    )
    def post(self, request):
        input_serializer = CalculationInputSerializer(data=request.data)

        if not input_serializer.is_valid():
            return Response(
                {'errors': input_serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        data = input_serializer.validated_data

        try:
            calculator = OzonCalculator(
                category_id=data['category_id'],
                price=data['price'],
                weight=data['weight'],
                volume=data['volume'],
                tax_rate=data['tax_rate'],
                buyout_rate=data['buyout_rate'],
                delivery_time=data['delivery_time'],
                ad_costs_rate=data['ad_costs_rate'],
                cost_price=data['cost_price'],
                other_costs=data['other_costs'],
                monthly_sales=data['monthly_sales']
            )
            results = calculator.calculate_all()
        except Category.DoesNotExist:
            return Response(
                {'error': f'Категория с ID {data["category_id"]} не найдена'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Ошибка при расчете: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        def serialize(value):
            if isinstance(value, Decimal):
                return float(value)
            return value

        fbo = results['fbo_results']
        fbs = results['fbs_results']

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = 'Итоги'

        summary_sheet.append(['Показатель', 'FBO', 'FBS'])
        metrics = [
            ('Цена', 'price'),
            ('Вознаграждение Ozon', 'ozon_reward'),
            ('Эквайринг', 'acquiring'),
            ('Обработка и доставка', 'processing_delivery'),
            ('Возвраты и отмены', 'returns_cancellations'),
            ('Затраты Ozon всего', 'total_ozon_costs'),
            ('Прибыль до собственных затрат', 'profit_before_costs'),
            ('Себестоимость', 'cost_price'),
            ('Налог на прибыль', 'profit_tax'),
            ('Прочие затраты', 'other_costs'),
            ('Чистая прибыль за шт', 'net_profit_per_unit'),
            ('Прибыль за месяц', 'net_profit_total'),
            ('Прибыль за год', 'annual_net_profit'),
        ]

        for label, key in metrics:
            summary_sheet.append([
                label,
                serialize(fbo.get(key, '')),
                serialize(fbs.get(key, '')),
            ])

        summary_sheet.append([])
        summary_sheet.append(['Показатель', 'FBO %', 'FBS %'])
        percent_metrics = [
            ('Эффективная комиссия Ozon', 'effective_ozon_fee_percent'),
            ('Валовая до налога', 'gross_margin_before_tax_percent'),
            ('Маржа на шт', 'net_profit_per_unit_percent'),
        ]
        for label, key in percent_metrics:
            summary_sheet.append([
                label,
                serialize(fbo.get(key, '')),
                serialize(fbs.get(key, '')),
            ])

        price_sheet = workbook.create_sheet('Чувствительность FBO')
        price_sheet.append(['Δ% цены', 'Цена', 'Прибыль/шт', 'Маржа %'])
        for row in fbo.get('sensitivity', {}).get('price', []):
            price_sheet.append([
                row.get('delta_pct'),
                serialize(row.get('price', '')),
                serialize(row.get('net_profit_per_unit', '')),
                serialize(row.get('net_profit_per_unit_percent', '')),
            ])

        buyout_sheet = workbook.create_sheet('Выкуп FBO')
        buyout_sheet.append(['Выкуп %', 'Прибыль/шт', 'Маржа %'])
        for row in fbo.get('sensitivity', {}).get('buyout', []):
            buyout_sheet.append([
                row.get('buyout_rate'),
                serialize(row.get('net_profit_per_unit', '')),
                serialize(row.get('net_profit_per_unit_percent', '')),
            ])

        delivery_sheet = workbook.create_sheet('Доставка FBO')
        delivery_sheet.append(['Часы', 'Прибыль/шт', 'Маржа %'])
        for row in fbo.get('sensitivity', {}).get('delivery_time', []):
            delivery_sheet.append([
                row.get('hours'),
                serialize(row.get('net_profit_per_unit', '')),
                serialize(row.get('net_profit_per_unit_percent', '')),
            ])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'ozon_calculation_{timestamp}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
