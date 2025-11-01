"""
Management команда для импорта категорий из официальной таблицы Ozon.

Использование:
    python manage.py import_ozon_categories путь/к/файлу.xlsx

Формат Excel файла (лист "Прайс (БЗ)"):
    - Колонка A (1): Категория (category_group)
    - Колонка B (2): Тип товара (name)
    - Колонка E (5): FBO комиссия (свыше 300 до 500 руб)
    - Колонка K (11): FBS комиссия (свыше 300 руб)
"""

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from categories.models import Category
import openpyxl
import os
from decimal import Decimal, InvalidOperation


class Command(BaseCommand):
    help = 'Импортирует категории товаров из официальной таблицы Ozon'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Путь к Excel файлу с категориями Ozon'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Прайс (БЗ)',
            help='Название листа в Excel (по умолчанию: "Прайс (БЗ)")'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Обновить существующие категории вместо пропуска',
            default=True
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Очистить все существующие категории перед импортом',
            default=False
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        sheet_name = options['sheet']
        update = options['update']
        clear = options['clear']

        # Проверка существования файла
        if not os.path.exists(excel_file):
            raise CommandError(f'Файл не найден: {excel_file}')

        # Открытие Excel файла
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            raise CommandError(f'Ошибка при открытии файла: {str(e)}')

        # Проверка наличия листа
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f'Лист "{sheet_name}" не найден. '
                f'Доступные листы: {", ".join(workbook.sheetnames)}'
            )

        worksheet = workbook[sheet_name]

        # Подсчет строк
        total_rows = worksheet.max_row
        if total_rows < 2:
            raise CommandError('Файл не содержит данных для импорта')

        self.stdout.write(f'Начинаю импорт из файла: {excel_file}')
        self.stdout.write(f'Лист: {sheet_name}')
        self.stdout.write(f'Найдено строк для обработки: {total_rows - 1}')

        # Очистка базы если указано
        if clear:
            deleted_count = Category.objects.all().count()
            Category.objects.all().delete()
            self.stdout.write(
                self.style.WARNING(f'Удалено существующих категорий: {deleted_count}')
            )

        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0

        # Импорт данных
        with transaction.atomic():
            # Пропускаем заголовок (строка 1)
            for row_num in range(2, total_rows + 1):
                row = worksheet[row_num]
                
                # Извлечение данных из ячеек
                # Колонка A (индекс 0): Категория
                category_group_cell = row[0].value if len(row) > 0 else None
                # Колонка B (индекс 1): Тип товара
                product_type_cell = row[1].value if len(row) > 1 else None
                # Колонка E (индекс 4): FBO комиссия (300-500 руб)
                fbo_cell = row[4].value if len(row) > 4 else None
                # Колонка K (индекс 10): FBS комиссия (свыше 300 руб)
                fbs_cell = row[10].value if len(row) > 10 else None

                # Пропуск пустых строк
                if not product_type_cell:
                    skipped_count += 1
                    continue

                # Преобразование названий в строки
                product_type = str(product_type_cell).strip()
                category_group = str(category_group_cell).strip() if category_group_cell else None
                
                if not product_type:
                    skipped_count += 1
                    continue

                # Преобразование комиссий
                try:
                    # FBO комиссия
                    if isinstance(fbo_cell, (int, float)):
                        fbo_commission = Decimal(str(fbo_cell)) * 100  # Преобразуем доли в проценты
                    elif fbo_cell:
                        fbo_commission = Decimal(str(fbo_cell).replace(',', '.')) * 100
                    else:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Строка {row_num}: отсутствует комиссия FBO для "{product_type}"'
                            )
                        )
                        error_count += 1
                        continue

                    # FBS комиссия
                    if isinstance(fbs_cell, (int, float)):
                        fbs_commission = Decimal(str(fbs_cell)) * 100
                    elif fbs_cell:
                        fbs_commission = Decimal(str(fbs_cell).replace(',', '.')) * 100
                    else:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Строка {row_num}: отсутствует комиссия FBS для "{product_type}"'
                            )
                        )
                        error_count += 1
                        continue

                    # Валидация значений комиссий
                    if fbo_commission < 0 or fbo_commission > 100:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Строка {row_num}: некорректная комиссия FBO ({fbo_commission}%) для "{product_type}"'
                            )
                        )
                        error_count += 1
                        continue

                    if fbs_commission < 0 or fbs_commission > 100:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Строка {row_num}: некорректная комиссия FBS ({fbs_commission}%) для "{product_type}"'
                            )
                        )
                        error_count += 1
                        continue

                except (InvalidOperation, ValueError, TypeError) as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f'Строка {row_num}: ошибка преобразования комиссий для "{product_type}": {str(e)}'
                        )
                    )
                    error_count += 1
                    continue

                # Создание или обновление категории
                try:
                    category, created = Category.objects.update_or_create(
                        name=product_type,
                        defaults={
                            'category_group': category_group,
                            'fbo_commission': fbo_commission,
                            'fbs_commission': fbs_commission,
                        }
                    )

                    if created:
                        created_count += 1
                        if created_count % 100 == 0:  # Прогресс каждые 100 записей
                            self.stdout.write(f'  Создано записей: {created_count}...')
                    elif update:
                        updated_count += 1
                        if updated_count % 100 == 0:
                            self.stdout.write(f'  Обновлено записей: {updated_count}...')
                    else:
                        skipped_count += 1

                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f'Строка {row_num}: ошибка при создании категории "{product_type}": {str(e)}'
                        )
                    )
                    error_count += 1
                    continue

        # Итоговая статистика
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write(self.style.SUCCESS('Импорт завершен!'))
        self.stdout.write(f'  Создано: {created_count}')
        if update:
            self.stdout.write(f'  Обновлено: {updated_count}')
        self.stdout.write(f'  Пропущено: {skipped_count}')
        if error_count > 0:
            self.stdout.write(self.style.ERROR(f'  Ошибок: {error_count}'))
        self.stdout.write('=' * 60)
        
        # Статистика по категориям
        total_categories = Category.objects.count()
        with_groups = Category.objects.exclude(category_group__isnull=True).count()
        self.stdout.write(f'\nВсего категорий в базе: {total_categories}')
        self.stdout.write(f'Категорий с группами: {with_groups}')







